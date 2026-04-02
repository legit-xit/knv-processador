from flask import Flask, request, send_file, render_template_string
import pandas as pd
import re
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import WriteOnlyCell

app = Flask(__name__)

HTML = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>KNV Processador</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            background: #0a0a0a;
            color: #fff;
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .container {
            max-width: 480px;
            width: 90%;
            text-align: center;
        }
        h1 {
            font-size: 1.8rem;
            margin-bottom: 0.5rem;
            background: linear-gradient(135deg, #6366f1, #8b5cf6);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        .sub { color: #666; margin-bottom: 2rem; font-size: 0.9rem; }
        .upload-area {
            border: 2px dashed #333;
            border-radius: 16px;
            padding: 3rem 2rem;
            cursor: pointer;
            transition: all 0.3s;
            background: #111;
        }
        .upload-area:hover, .upload-area.dragover {
            border-color: #6366f1;
            background: #1a1a2e;
        }
        .upload-area p { color: #888; margin-top: 0.5rem; font-size: 0.85rem; }
        .upload-icon { font-size: 2.5rem; margin-bottom: 0.5rem; }
        input[type="file"] { display: none; }
        .btn {
            display: none;
            width: 100%;
            padding: 1rem;
            margin-top: 1.5rem;
            border: none;
            border-radius: 12px;
            background: linear-gradient(135deg, #6366f1, #8b5cf6);
            color: #fff;
            font-size: 1rem;
            font-weight: 600;
            cursor: pointer;
            transition: opacity 0.3s;
        }
        .btn:hover { opacity: 0.9; }
        .btn:disabled { opacity: 0.5; cursor: wait; }
        .file-name {
            margin-top: 1rem;
            padding: 0.75rem 1rem;
            background: #1a1a2e;
            border-radius: 8px;
            font-size: 0.85rem;
            color: #a5b4fc;
            display: none;
        }
        .status {
            margin-top: 1rem;
            font-size: 0.85rem;
            color: #22c55e;
            display: none;
        }
        .error { color: #ef4444; }
    </style>
</head>
<body>
    <div class="container">
        <h1>KNV Processador</h1>
        <p class="sub">Joga o CSV bruto e baixa o .xlsx pronto</p>

        <div class="upload-area" id="dropZone" onclick="document.getElementById('fileInput').click()">
            <div class="upload-icon">📄</div>
            <strong>Clica ou arrasta o CSV aqui</strong>
            <p>Arquivo .csv da plataforma</p>
        </div>

        <div class="file-name" id="fileName"></div>
        <input type="file" id="fileInput" accept=".csv">
        <button class="btn" id="processBtn" onclick="processar()">Processar e Baixar</button>
        <div class="status" id="status"></div>
    </div>

    <script>
        const dropZone = document.getElementById('dropZone');
        const fileInput = document.getElementById('fileInput');
        const fileName = document.getElementById('fileName');
        const processBtn = document.getElementById('processBtn');
        const status = document.getElementById('status');

        dropZone.addEventListener('dragover', e => {
            e.preventDefault();
            dropZone.classList.add('dragover');
        });
        dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
        dropZone.addEventListener('drop', e => {
            e.preventDefault();
            dropZone.classList.remove('dragover');
            if (e.dataTransfer.files[0]) {
                fileInput.files = e.dataTransfer.files;
                mostrarArquivo(e.dataTransfer.files[0]);
            }
        });
        fileInput.addEventListener('change', () => {
            if (fileInput.files[0]) mostrarArquivo(fileInput.files[0]);
        });

        function mostrarArquivo(file) {
            fileName.textContent = file.name + ' (' + (file.size / 1024 / 1024).toFixed(1) + ' MB)';
            fileName.style.display = 'block';
            processBtn.style.display = 'block';
            status.style.display = 'none';
        }

        async function processar() {
            const file = fileInput.files[0];
            if (!file) return;

            processBtn.disabled = true;
            processBtn.textContent = 'Processando...';
            status.style.display = 'none';

            const formData = new FormData();
            formData.append('file', file);

            try {
                const res = await fetch('/processar', { method: 'POST', body: formData });
                if (!res.ok) {
                    const err = await res.text();
                    throw new Error(err);
                }
                const blob = await res.blob();
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = res.headers.get('X-Filename') || 'KNV.xlsx';
                a.click();
                URL.revokeObjectURL(url);

                status.textContent = '✅ Pronto! Download iniciado.';
                status.className = 'status';
                status.style.display = 'block';
            } catch (err) {
                status.textContent = '❌ Erro: ' + err.message;
                status.className = 'status error';
                status.style.display = 'block';
            }

            processBtn.disabled = false;
            processBtn.textContent = 'Processar e Baixar';
        }
    </script>
</body>
</html>
"""


@app.route('/')
def index():
    return render_template_string(HTML)


@app.route('/processar', methods=['POST'])
def processar():
    if 'file' not in request.files:
        return 'Nenhum arquivo enviado', 400

    file = request.files['file']

    try:
        df = pd.read_csv(file, sep=';', encoding='latin1', dtype=str, low_memory=False)
    except Exception:
        try:
            file.seek(0)
            df = pd.read_csv(file, sep=';', encoding='utf-8', dtype=str, low_memory=False)
        except Exception as e:
            return f'Erro ao ler CSV: {e}', 400

    def clean(val):
        if pd.isna(val):
            return ''
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', str(val))

    df = df.map(clean)

    cols_to_drop = [
        'Comissão', 'Desconto (Valor)', 'Desconto (Automático)', 'Taxas',
        'Parcelamento sem juros', 'Cliente (IP)', 'Cliente (CEP)',
        'Cliente (Logradouro)', 'Cliente (Número)', 'Cliente (Complemento)',
        'Cliente (Bairro)', 'Cliente (Cidade)', 'Cliente (Estado)',
        'Cliente (País)', 'Afiliado (Nome)', 'Afiliado (E-mail)',
        'UTM SRC', 'UTM Source', 'UTM Medium', 'UTM Campaign',
        'UTM Term', 'UTM Content'
    ]
    df = df.drop(columns=[c for c in cols_to_drop if c in df.columns])

    today = date.today()
    filename = f"kirvano_{today.day:02d}-{today.month:02d}-{today.year}.xlsx"

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_align = Alignment(vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    green = PatternFill('solid', fgColor='C6EFCE')
    red = PatternFill('solid', fgColor='FFC7CE')

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Vendas")

    headers = list(df.columns)
    status_idx = headers.index('Status') if 'Status' in headers else None

    h_cells = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border
        h_cells.append(c)
    ws.append(h_cells)

    for row in df.itertuples(index=False):
        fill = None
        if status_idx is not None:
            status = str(row[status_idx]).strip().lower()
            if status == 'aprovada':
                fill = green
            elif status in ('chargeback', 'med', 'estornada'):
                fill = red

        cells = []
        for val in row:
            c = WriteOnlyCell(ws, value=val)
            c.font = cell_font
            c.alignment = cell_align
            c.border = thin_border
            if fill:
                c.fill = fill
            cells.append(c)
        ws.append(cells)

    output_path = f'/tmp/{filename}'
    wb.save(output_path)

    response = send_file(output_path, as_attachment=True, download_name=filename)
    response.headers['X-Filename'] = filename
    return response


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
